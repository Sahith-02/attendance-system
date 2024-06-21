from flask import Flask, render_template, request, redirect, url_for
from datetime import date, datetime, timedelta
from openpyxl import load_workbook

app = Flask(__name__)

# Dummy user database (replace with a proper database in production)
users = {
    'user1': {
        'username': 'user1',
        'password': 'password1'
    }
}

# Routes

@app.route('/')
def index():
    return render_template('home.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username in users and users[username]['password'] == password:
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error='Invalid username or password')
    return render_template('login.html')



users = {}

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # Add user to the dictionary (replace with database operations as needed)
        users[username] = {'username': username, 'password': password}
        
        # Redirect to login page after successful registration
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/current_day')
def current_day():
    # Load data from Excel sheet for current day
    workbook = load_workbook('data/current_day.xlsx')
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append({'time': row[1], 'name': row[2]})
    return render_template('attendance.html', title='Current Day Attendance', data=data)

def calculate_active_days():
    today = datetime.now()
    first_day_of_month = today.replace(day=1)
    next_month = today.replace(month=today.month % 12 + 1, day=1)
    last_day_of_month = next_month - timedelta(days=1)

    current_date = first_day_of_month
    active_days = 0

    while current_date <= last_day_of_month:
        if current_date.weekday() < 5:  # Monday to Friday (0 to 4)
            active_days += 1
        current_date += timedelta(days=1)

    return active_days

@app.route('/current_month')
def current_month():
    # Load data from Excel sheet for current month
    workbook = load_workbook('data/current_month.xlsx')
    sheet = workbook.active

    # Calculate active days and initialize variables
    active_days = calculate_active_days()
    data = []

    # Iterate through rows starting from row 2 (assuming row 1 is header)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        days_present = row[0]  # Assuming days present is in column A
        name = row[1]           # Assuming name is in column B
        
        if active_days > 0:
            attendance_percentage = (days_present / active_days) * 100
        else:
            attendance_percentage = 0

        # Round percentage to two decimal places
        attendance_percentage = round(attendance_percentage, 2)

        data.append({'name': name, 'attendance_percentage': attendance_percentage})

    return render_template('current_month.html', title='Current Month Attendance', data=data)


if __name__ == '__main__':
    app.run(debug=True)
